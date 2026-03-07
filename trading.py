import os
import random
import re
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from functools import lru_cache
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from flask import has_app_context
from openpyxl import load_workbook
from sqlalchemy.exc import OperationalError, ProgrammingError

from models import CFDSymbol


DEFAULT_CFD_SYMBOL_SPECS = (
    {"symbol": "EURUSD", "aliases": ("EU",), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 10},
    {"symbol": "GBPUSD", "aliases": ("GU",), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 20},
    {"symbol": "USDJPY", "aliases": ("UJ",), "contract_size": 100000.0, "pip_size": 0.01, "sort_order": 30},
    {"symbol": "AUDUSD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 40},
    {"symbol": "USDCAD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 50},
    {"symbol": "USDCHF", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 60},
    {"symbol": "NZDUSD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 70},
    {"symbol": "EURJPY", "aliases": (), "contract_size": 100000.0, "pip_size": 0.01, "sort_order": 80},
    {"symbol": "GBPJPY", "aliases": (), "contract_size": 100000.0, "pip_size": 0.01, "sort_order": 90},
    {"symbol": "AUDJPY", "aliases": (), "contract_size": 100000.0, "pip_size": 0.01, "sort_order": 100},
    {"symbol": "CADJPY", "aliases": (), "contract_size": 100000.0, "pip_size": 0.01, "sort_order": 110},
    {"symbol": "CHFJPY", "aliases": (), "contract_size": 100000.0, "pip_size": 0.01, "sort_order": 120},
    {"symbol": "NZDJPY", "aliases": (), "contract_size": 100000.0, "pip_size": 0.01, "sort_order": 130},
    {"symbol": "EURGBP", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 140},
    {"symbol": "EURCHF", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 150},
    {"symbol": "EURAUD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 160},
    {"symbol": "EURNZD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 170},
    {"symbol": "EURCAD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 180},
    {"symbol": "GBPCHF", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 190},
    {"symbol": "GBPAUD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 200},
    {"symbol": "GBPCAD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 210},
    {"symbol": "GBPNZD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 220},
    {"symbol": "AUDCAD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 230},
    {"symbol": "AUDCHF", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 240},
    {"symbol": "AUDNZD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 250},
    {"symbol": "CADCHF", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 260},
    {"symbol": "NZDCHF", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 270},
    {"symbol": "NZDCAD", "aliases": (), "contract_size": 100000.0, "pip_size": 0.0001, "sort_order": 280},
    {"symbol": "XAUUSD", "aliases": (), "contract_size": 100.0, "pip_size": None, "sort_order": 290},
    {"symbol": "XAGUSD", "aliases": (), "contract_size": 5000.0, "pip_size": None, "sort_order": 300},
    {"symbol": "US500", "aliases": ("SPX500", "SP500", "US500CASH", "US500INDEX"), "contract_size": 1.0, "pip_size": None, "sort_order": 310},
    {"symbol": "NAS100", "aliases": ("US100", "USTEC", "NAS100CASH", "NASDAQ100"), "contract_size": 1.0, "pip_size": None, "sort_order": 320},
    {"symbol": "US30", "aliases": ("DJ30", "DJIA", "WS30", "US30CASH"), "contract_size": 1.0, "pip_size": None, "sort_order": 330},
    {"symbol": "GER40", "aliases": ("DE40", "DAX40", "GER30", "DE30", "DAX"), "contract_size": 1.0, "pip_size": None, "sort_order": 340},
    {"symbol": "UK100", "aliases": ("FTSE100", "UKX", "U100"), "contract_size": 1.0, "pip_size": None, "sort_order": 350},
    {"symbol": "FRA40", "aliases": ("CAC40", "FR40"), "contract_size": 1.0, "pip_size": None, "sort_order": 360},
    {"symbol": "EU50", "aliases": ("STOXX50", "EUSTX50", "SX5E"), "contract_size": 1.0, "pip_size": None, "sort_order": 370},
    {"symbol": "JP225", "aliases": ("N225", "NI225", "JP225CASH"), "contract_size": 1.0, "pip_size": None, "sort_order": 380},
    {"symbol": "HK50", "aliases": ("HSI", "HSI50", "HK50CASH"), "contract_size": 1.0, "pip_size": None, "sort_order": 390},
    {"symbol": "CHN50", "aliases": ("CN50", "CHINA50", "CHINAA50", "A50"), "contract_size": 1.0, "pip_size": None, "sort_order": 400},
    {"symbol": "AUS200", "aliases": ("AU200", "ASX200"), "contract_size": 1.0, "pip_size": None, "sort_order": 410},
    {"symbol": "ESP35", "aliases": ("IBEX35", "ES35"), "contract_size": 1.0, "pip_size": None, "sort_order": 420},
    {"symbol": "IT40", "aliases": ("ITA40",), "contract_size": 1.0, "pip_size": None, "sort_order": 430},
)

MT5_COLUMN_ALIASES = {
    "mt5_position": {"position", "position id", "position ticket", "ticket"},
    "symbol": {"symbol", "instrument"},
    "side": {"type", "deal", "side"},
    "lot_size": {"volume", "lots", "lot"},
    "entry_price": {"open price", "price", "entry price"},
    "exit_price": {"close price", "exit price", "price close"},
    "pnl": {"profit", "p/l", "pl", "net profit"},
    "opened_at": {"time", "date", "open time", "opened at", "time open"},
    "closed_at": {"close time", "closed at", "time close", "closing time", "closed time"},
}

MT5_SECTION_TITLES = {"positions", "orders", "deals", "results"}

WEEKDAY_NAMES = [
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
]

SESSION_DEFINITIONS = (
    {"name": "Sydney", "zone": "Australia/Sydney", "start_hour": 7, "end_hour": 16},
    {"name": "Tokyo", "zone": "Asia/Tokyo", "start_hour": 9, "end_hour": 18},
    {"name": "London", "zone": "Europe/London", "start_hour": 8, "end_hour": 17},
    {"name": "New York", "zone": "America/New_York", "start_hour": 8, "end_hour": 17},
)


def normalize_symbol(symbol):
    return "".join(ch for ch in (symbol or "").upper() if ch.isalnum())


def _normalize_aliases(value):
    seen = []
    for raw_alias in str(value or "").split(","):
        alias = normalize_symbol(raw_alias)
        if alias and alias not in seen:
            seen.append(alias)
    return tuple(seen)


def clear_cfd_symbol_cache():
    _load_cfd_symbol_specs.cache_clear()
    _load_cfd_symbol_map.cache_clear()
    _load_cfd_alias_map.cache_clear()


@lru_cache(maxsize=1)
def _load_cfd_symbol_specs():
    default_specs = tuple(
        {
            "symbol": spec["symbol"],
            "aliases": tuple(spec["aliases"]),
            "contract_size": float(spec["contract_size"]),
            "pip_size": spec["pip_size"],
            "sort_order": int(spec["sort_order"]),
        }
        for spec in DEFAULT_CFD_SYMBOL_SPECS
    )
    if not has_app_context():
        return default_specs

    try:
        rows = (
            CFDSymbol.query.filter_by(is_active=True)
            .order_by(CFDSymbol.sort_order.asc(), CFDSymbol.symbol.asc())
            .all()
        )
    except (OperationalError, ProgrammingError):
        return default_specs

    if not rows:
        return default_specs

    return tuple(
        {
            "symbol": normalize_symbol(row.symbol),
            "aliases": _normalize_aliases(row.aliases),
            "contract_size": float(row.contract_size),
            "pip_size": float(row.pip_size) if row.pip_size is not None else None,
            "sort_order": int(row.sort_order or 0),
        }
        for row in rows
    )


@lru_cache(maxsize=1)
def _load_cfd_symbol_map():
    return {spec["symbol"]: spec for spec in _load_cfd_symbol_specs()}


@lru_cache(maxsize=1)
def _load_cfd_alias_map():
    alias_map = {}
    for spec in _load_cfd_symbol_specs():
        alias_map[spec["symbol"]] = spec["symbol"]
        for alias in spec["aliases"]:
            alias_map[alias] = spec["symbol"]
    return alias_map


def canonicalize_symbol(symbol):
    normalized = normalize_symbol(symbol)
    if not normalized:
        return ""
    return _load_cfd_alias_map().get(normalized, normalized)


def get_symbol_options(selected_symbol=None):
    options = [spec["symbol"] for spec in _load_cfd_symbol_specs()]
    selected = canonicalize_symbol(selected_symbol)
    if selected and selected not in options:
        options.insert(0, selected)
    return options


def normalize_header_name(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s_/\\-]+", " ", text)
    text = re.sub(r"[^a-z0-9 ]", "", text)
    return " ".join(text.split())


def parse_side_value(value):
    text = str(value or "").strip().upper()
    if "BUY" in text:
        return "BUY"
    if "SELL" in text:
        return "SELL"
    return None


def parse_float_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    try:
        return float(text)
    except ValueError:
        return None


def parse_datetime_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y.%m.%d %H:%M:%S",
        "%Y.%m.%d %H:%M",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_mt5_position_value(value):
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return None

    text = str(value).strip()
    if not text:
        return None

    cleaned = text.replace(",", "")
    if re.fullmatch(r"\d+", cleaned):
        return cleaned
    if re.fullmatch(r"\d+\.0+", cleaned):
        return cleaned.split(".", 1)[0]
    return None


def build_import_signature(prefix="mt5"):
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    random_tail = os.urandom(4).hex()
    return f"{prefix}_{timestamp}_{random_tail}"


def parse_import_signature_datetime(signature):
    text_value = str(signature or "").strip()
    match = re.match(r"^[A-Za-z0-9]+_(\d{14})_[0-9a-fA-F]+$", text_value)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%Y%m%d%H%M%S")
    except ValueError:
        return None


def row_texts(row):
    return [str(cell).strip() for cell in row if cell is not None and str(cell).strip() != ""]


def find_section_row(rows, section_name):
    target = normalize_header_name(section_name)
    for idx, row in enumerate(rows):
        texts = row_texts(row)
        if len(texts) == 1 and normalize_header_name(texts[0]) == target:
            return idx
    return None


def is_section_title_row(row):
    texts = row_texts(row)
    if len(texts) != 1:
        return False
    return normalize_header_name(texts[0]) in MT5_SECTION_TITLES


def build_mt5_column_map(header_row):
    normalized = [normalize_header_name(cell) for cell in header_row]
    column_map = {}
    by_name = {}

    for idx, name in enumerate(normalized):
        if not name:
            continue
        by_name.setdefault(name, []).append(idx)
        for canonical, aliases in MT5_COLUMN_ALIASES.items():
            if canonical in column_map:
                continue
            if name in aliases:
                column_map[canonical] = idx
                break

    time_cols = by_name.get("time", [])
    price_cols = by_name.get("price", [])
    if "opened_at" not in column_map and time_cols:
        column_map["opened_at"] = time_cols[0]
    if "closed_at" not in column_map and len(time_cols) >= 2:
        column_map["closed_at"] = time_cols[1]
    if "entry_price" not in column_map and price_cols:
        column_map["entry_price"] = price_cols[0]
    if "exit_price" not in column_map and len(price_cols) >= 2:
        column_map["exit_price"] = price_cols[1]

    return column_map


def parse_mt5_xlsx_stream(file_stream):
    workbook = load_workbook(file_stream, data_only=True, read_only=True)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    section_idx = find_section_row(rows, "positions")
    if section_idx is None:
        workbook.close()
        return [], 0, 0

    header_idx = None
    column_map = {}

    header_scan_end = min(section_idx + 10, len(rows))
    for idx in range(section_idx + 1, header_scan_end):
        detected = build_mt5_column_map(rows[idx])
        if (
            detected.get("symbol") is not None
            and detected.get("side") is not None
            and detected.get("lot_size") is not None
        ):
            header_idx = idx
            column_map = detected
            break

    if header_idx is None:
        workbook.close()
        return [], 0, 0

    parsed = []
    skipped = 0
    total = 0

    for row_idx in range(header_idx + 1, len(rows)):
        row = rows[row_idx]
        texts = row_texts(row)
        if not texts:
            continue
        if is_section_title_row(row):
            break

        total += 1

        def col(key):
            idx = column_map.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        symbol = canonicalize_symbol(col("symbol"))
        mt5_position_raw = col("mt5_position")
        mt5_position = parse_mt5_position_value(mt5_position_raw)
        side = parse_side_value(col("side"))
        lot_size = parse_float_value(col("lot_size"))
        entry_price = parse_float_value(col("entry_price"))
        exit_price = parse_float_value(col("exit_price"))
        pnl = parse_float_value(col("pnl"))
        opened_at = parse_datetime_value(col("opened_at"))
        closed_at = parse_datetime_value(col("closed_at"))

        if not symbol or not side or lot_size is None or entry_price is None:
            skipped += 1
            continue

        parsed.append(
            {
                "symbol": symbol,
                "mt5_position": mt5_position,
                "mt5_position_raw": mt5_position_raw,
                "side": side,
                "lot_size": lot_size,
                "entry_price": entry_price,
                "exit_price": exit_price,
                "pnl": pnl,
                "opened_at": opened_at,
                "closed_at": closed_at,
            }
        )

    workbook.close()
    return parsed, total, skipped


def is_fx_pair(symbol):
    sym = normalize_symbol(symbol)
    return len(sym) == 6 and sym.isalpha()


def get_contract_size(symbol):
    sym = canonicalize_symbol(symbol)
    spec = _load_cfd_symbol_map().get(sym)
    if spec is not None:
        return spec["contract_size"]
    if is_fx_pair(sym):
        return 100000.0
    return 1.0


def quote_to_usd_rate(symbol, reference_price):
    sym = normalize_symbol(symbol)
    if sym.endswith("USD"):
        return 1.0
    if is_fx_pair(sym) and sym.startswith("USD") and reference_price and reference_price > 0:
        return 1.0 / reference_price
    return None


def get_pip_size(symbol):
    sym = canonicalize_symbol(symbol)
    spec = _load_cfd_symbol_map().get(sym)
    if spec is not None and spec["pip_size"] is not None:
        return spec["pip_size"]
    if not is_fx_pair(sym):
        return None
    return 0.01 if sym.endswith("JPY") else 0.0001


def calc_pips_values(symbol, side, entry_price, exit_price):
    if entry_price is None or exit_price is None:
        return None

    pip_size = get_pip_size(symbol)
    if not pip_size:
        return None

    direction = 1.0 if (side or "").upper() == "BUY" else -1.0
    return direction * ((exit_price - entry_price) / pip_size)


def calc_pnl_values(symbol, side, entry_price, exit_price, lot_size):
    if entry_price is None or exit_price is None:
        return None
    if lot_size is None or lot_size <= 0:
        return None

    direction = 1.0 if (side or "").upper() == "BUY" else -1.0
    contract_size = get_contract_size(symbol)
    conversion_rate = quote_to_usd_rate(symbol, exit_price)
    if conversion_rate is None:
        return None

    return direction * (exit_price - entry_price) * lot_size * contract_size * conversion_rate


def derive_exit_price(symbol, side, entry_price, lot_size, pnl_value):
    if entry_price is None or pnl_value is None:
        return None
    if lot_size is None or lot_size <= 0:
        return None

    direction = 1.0 if (side or "").upper() == "BUY" else -1.0
    contract_size = get_contract_size(symbol)
    units = lot_size * contract_size
    if units <= 0:
        return None

    sym = normalize_symbol(symbol)
    if sym.endswith("USD"):
        return entry_price + (pnl_value / (direction * units))

    if is_fx_pair(sym) and sym.startswith("USD"):
        k = pnl_value / (direction * units)
        denom = 1.0 - k
        if abs(denom) < 1e-9:
            return None
        candidate = entry_price / denom
        return candidate if candidate > 0 else None

    return None


def calc_pnl(trade):
    return calc_pnl_values(
        symbol=trade.symbol,
        side=trade.side,
        entry_price=trade.entry_price,
        exit_price=trade.exit_price,
        lot_size=trade.lot_size,
    )


def resolve_pnl(trade):
    if trade.pnl is not None:
        return trade.pnl
    return calc_pnl(trade)


def resolve_pips(trade):
    return calc_pips_values(
        symbol=trade.symbol,
        side=trade.side,
        entry_price=trade.entry_price,
        exit_price=trade.exit_price,
    )


def build_dashboard_insight(user_trades, closed_trades):
    total_trades = len(user_trades)
    total_closed = len(closed_trades)
    insights = []

    symbol_counts = {}
    for trade in user_trades:
        symbol = (trade.symbol or "").strip().upper()
        if not symbol:
            continue
        symbol_counts[symbol] = symbol_counts.get(symbol, 0) + 1
    if symbol_counts:
        top_symbol, top_count = max(symbol_counts.items(), key=lambda item: (item[1], item[0]))
        top_share = (top_count / max(sum(symbol_counts.values()), 1)) * 100
        insights.append(
            {
                "title": "Most Traded Pair",
                "body": (
                    f"{top_symbol} is your most traded pair with {top_count} trades "
                    f"({top_share:.0f}% of your journal)."
                ),
            }
        )

    if total_closed >= 3:
        total_wins = sum(1 for _, pnl in closed_trades if pnl > 0)
        overall_win_rate = (total_wins / total_closed) * 100

        pair_stats = {}
        for trade, pnl in closed_trades:
            symbol = (trade.symbol or "").strip().upper()
            if not symbol:
                continue
            if symbol not in pair_stats:
                pair_stats[symbol] = {"count": 0, "wins": 0, "pnl_sum": 0.0}
            pair_stats[symbol]["count"] += 1
            pair_stats[symbol]["pnl_sum"] += float(pnl)
            if pnl > 0:
                pair_stats[symbol]["wins"] += 1

        eligible_pairs = [
            (symbol, stats) for symbol, stats in pair_stats.items() if stats["count"] >= 2
        ]
        if eligible_pairs:
            best_pair, best_stats = max(
                eligible_pairs,
                key=lambda item: (item[1]["wins"] / item[1]["count"], item[1]["count"]),
            )
            pair_win_rate = (best_stats["wins"] / best_stats["count"]) * 100
            pair_avg_pnl = best_stats["pnl_sum"] / best_stats["count"]
            if pair_win_rate > overall_win_rate:
                win_rate_delta = pair_win_rate - overall_win_rate
                insights.append(
                    {
                        "title": "Pair Edge",
                        "body": (
                            f"You perform {win_rate_delta:.1f}% better on {best_pair} by win rate "
                            f"({pair_win_rate:.1f}% vs {overall_win_rate:.1f}% overall)."
                        ),
                    }
                )
            insights.append(
                {
                    "title": "Pair Efficiency",
                    "body": (
                        f"Your average closed-trade result on {best_pair} is "
                        f"{pair_avg_pnl:+.2f} across {best_stats['count']} trades."
                    ),
                }
            )

    if total_closed >= 4:
        side_stats = {"BUY": {"count": 0, "pnl_sum": 0.0}, "SELL": {"count": 0, "pnl_sum": 0.0}}
        for trade, pnl in closed_trades:
            side = (trade.side or "").strip().upper()
            if side not in side_stats:
                continue
            side_stats[side]["count"] += 1
            side_stats[side]["pnl_sum"] += float(pnl)

        buy_count = side_stats["BUY"]["count"]
        sell_count = side_stats["SELL"]["count"]
        if buy_count >= 2 and sell_count >= 2:
            buy_avg = side_stats["BUY"]["pnl_sum"] / buy_count
            sell_avg = side_stats["SELL"]["pnl_sum"] / sell_count
            if abs(buy_avg - sell_avg) >= 0.01:
                better_side = "BUY" if buy_avg > sell_avg else "SELL"
                better_avg = buy_avg if buy_avg > sell_avg else sell_avg
                weaker_side = "SELL" if better_side == "BUY" else "BUY"
                weaker_avg = sell_avg if better_side == "BUY" else buy_avg
                insights.append(
                    {
                        "title": "Directional Bias",
                        "body": (
                            f"{better_side} setups outperform {weaker_side} "
                            f"({better_avg:+.2f} vs {weaker_avg:+.2f} average PnL per closed trade)."
                        ),
                    }
                )

    if total_closed >= 6:
        day_totals = {}
        for trade, pnl in closed_trades:
            if not trade.opened_at:
                continue
            day_key = trade.opened_at.strftime("%Y-%m-%d")
            if day_key not in day_totals:
                day_totals[day_key] = {"pnl_sum": 0.0, "count": 0}
            day_totals[day_key]["pnl_sum"] += float(pnl)
            day_totals[day_key]["count"] += 1

        if day_totals:
            trades_on_losing_days = sum(
                data["count"] for data in day_totals.values() if data["pnl_sum"] < 0
            )
            trades_on_non_losing_days = sum(
                data["count"] for data in day_totals.values() if data["pnl_sum"] >= 0
            )
            total_day_trades = trades_on_losing_days + trades_on_non_losing_days
            if total_day_trades > 0 and trades_on_losing_days > trades_on_non_losing_days:
                losing_day_ratio = (trades_on_losing_days / total_day_trades) * 100
                insights.append(
                    {
                        "title": "Risk Pattern",
                        "body": (
                            f"{losing_day_ratio:.0f}% of your trades happen on net-losing days. "
                            "Consider reducing frequency after early losses."
                        ),
                    }
                )

    if total_trades == 0:
        return {
            "title": "No Data Yet",
            "body": "Add your first trade to unlock personalized performance insights.",
        }
    if not insights:
        return {
            "title": "Developing Edge",
            "body": (
                f"You've logged {total_trades} trades so far. "
                "As more closed trades accumulate, your insights will become more specific."
            ),
        }
    return random.choice(insights)


def get_timezone(name):
    try:
        return ZoneInfo(str(name or "UTC").strip() or "UTC")
    except ZoneInfoNotFoundError:
        return timezone.utc


def ensure_utc_aware(value):
    if value is None:
        return None
    if value.tzinfo is not None:
        return value.astimezone(timezone.utc)
    return value.replace(tzinfo=timezone.utc)


def to_display_timezone(value, timezone_name):
    aware_value = ensure_utc_aware(value)
    if aware_value is None:
        return None
    return aware_value.astimezone(get_timezone(timezone_name))


def format_duration_minutes(duration_minutes):
    if duration_minutes is None:
        return "-"
    total_minutes = int(round(duration_minutes))
    hours, minutes = divmod(total_minutes, 60)
    days, hours = divmod(hours, 24)
    parts = []
    if days:
        parts.append(f"{days}d")
    if hours:
        parts.append(f"{hours}h")
    if minutes or not parts:
        parts.append(f"{minutes}m")
    return " ".join(parts)


def get_active_sessions(timestamp_utc):
    if timestamp_utc is None:
        return []

    active_sessions = []
    for session_def in SESSION_DEFINITIONS:
        session_zone = get_timezone(session_def["zone"])
        local_value = timestamp_utc.astimezone(session_zone)
        hour_value = local_value.hour + (local_value.minute / 60.0)
        if session_def["start_hour"] <= hour_value < session_def["end_hour"]:
            active_sessions.append(session_def["name"])
    return active_sessions


def classify_trading_session(timestamp_utc):
    active_sessions = get_active_sessions(timestamp_utc)
    if not active_sessions:
        return "Off Hours"
    if len(active_sessions) == 1:
        return active_sessions[0]
    return " / ".join(active_sessions)


def calculate_streaks(closed_records):
    if not closed_records:
        return {
            "current_type": "none",
            "current_length": 0,
            "best_win_streak": 0,
            "best_loss_streak": 0,
        }

    best_win_streak = 0
    best_loss_streak = 0
    current_type = "none"
    current_length = 0

    for record in closed_records:
        pnl_value = record["pnl"]
        if pnl_value > 0:
            streak_type = "win"
        elif pnl_value < 0:
            streak_type = "loss"
        else:
            streak_type = "breakeven"

        if streak_type == current_type:
            current_length += 1
        else:
            current_type = streak_type
            current_length = 1

        if streak_type == "win":
            best_win_streak = max(best_win_streak, current_length)
        elif streak_type == "loss":
            best_loss_streak = max(best_loss_streak, current_length)

    return {
        "current_type": current_type,
        "current_length": current_length,
        "best_win_streak": best_win_streak,
        "best_loss_streak": best_loss_streak,
    }


def build_trade_analytics(trades, display_timezone_name="UTC", now_utc=None):
    display_timezone = get_timezone(display_timezone_name)
    now_utc = ensure_utc_aware(now_utc or datetime.utcnow())
    now_local = now_utc.astimezone(display_timezone)
    week_start_local = (now_local - timedelta(days=now_local.weekday())).replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0,
    )
    month_start_local = now_local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    total_trades = len(trades)
    closed_records = []
    open_trades = 0

    for trade in sorted(trades, key=lambda item: ((item.opened_at or datetime.min), item.id or 0)):
        pnl_value = resolve_pnl(trade)
        opened_at_utc = ensure_utc_aware(trade.opened_at)
        closed_at_utc = ensure_utc_aware(getattr(trade, "closed_at", None))
        duration_minutes = None
        if opened_at_utc and closed_at_utc and closed_at_utc >= opened_at_utc:
            duration_minutes = (closed_at_utc - opened_at_utc).total_seconds() / 60.0

        if pnl_value is None:
            open_trades += 1
            continue

        opened_local = opened_at_utc.astimezone(display_timezone) if opened_at_utc else None
        closed_local = closed_at_utc.astimezone(display_timezone) if closed_at_utc else None
        record = {
            "trade": trade,
            "pnl": float(pnl_value),
            "pips": resolve_pips(trade),
            "symbol": (trade.symbol or "").strip().upper(),
            "side": (trade.side or "").strip().upper(),
            "opened_at_utc": opened_at_utc,
            "closed_at_utc": closed_at_utc,
            "opened_at_local": opened_local,
            "closed_at_local": closed_local,
            "opened_label": opened_local.strftime("%d %b %Y %H:%M") if opened_local else "-",
            "closed_label": closed_local.strftime("%d %b %Y %H:%M") if closed_local else "-",
            "weekday": WEEKDAY_NAMES[opened_local.weekday()] if opened_local else "Unknown",
            "session": classify_trading_session(opened_at_utc),
            "duration_minutes": duration_minutes,
            "duration_label": format_duration_minutes(duration_minutes),
            "status": "Closed" if trade.exit_price is not None else "Running",
        }
        closed_records.append(record)

    total_closed = len(closed_records)
    wins = sum(1 for record in closed_records if record["pnl"] > 0)
    losses = sum(1 for record in closed_records if record["pnl"] < 0)
    breakeven = total_closed - wins - losses
    gross_profit = sum(record["pnl"] for record in closed_records if record["pnl"] > 0)
    gross_loss = sum(record["pnl"] for record in closed_records if record["pnl"] < 0)
    net_pnl = sum(record["pnl"] for record in closed_records)
    avg_win = (gross_profit / wins) if wins else None
    avg_loss_abs = (abs(gross_loss) / losses) if losses else None
    win_rate = (wins / total_closed * 100.0) if total_closed else 0.0
    expectancy = (net_pnl / total_closed) if total_closed else None
    payoff_ratio = (avg_win / avg_loss_abs) if avg_win is not None and avg_loss_abs else None
    risk_reward_ratio = (avg_loss_abs / avg_win) if avg_loss_abs is not None and avg_win else None
    profit_factor = (gross_profit / abs(gross_loss)) if gross_loss < 0 else None

    duration_values = [
        record["duration_minutes"]
        for record in closed_records
        if record["duration_minutes"] is not None
    ]
    average_duration_minutes = (
        sum(duration_values) / len(duration_values) if duration_values else None
    )
    duration_coverage = (
        len(duration_values) / total_closed * 100.0 if total_closed else 0.0
    )

    weekly_pnl = 0.0
    monthly_pnl = 0.0
    trading_days = set()
    equity_curve = []
    running_equity = 0.0
    daily_equity = {}
    peak_equity = 0.0
    max_drawdown = 0.0

    for record in closed_records:
        opened_local = record["opened_at_local"]
        if opened_local:
            trading_days.add(opened_local.date().isoformat())
            if opened_local >= week_start_local:
                weekly_pnl += record["pnl"]
            if opened_local >= month_start_local:
                monthly_pnl += record["pnl"]

        running_equity += record["pnl"]
        peak_equity = max(peak_equity, running_equity)
        max_drawdown = min(max_drawdown, running_equity - peak_equity)
        record["equity_after_trade"] = round(running_equity, 2)

        if opened_local:
            day_key = opened_local.date().isoformat()
            daily_equity[day_key] = {
                "date": day_key,
                "label": opened_local.strftime("%d %b"),
                "equity": round(running_equity, 2),
            }

        equity_curve.append(
            {
                "date": record["opened_at_local"].strftime("%Y-%m-%d %H:%M")
                if record["opened_at_local"]
                else f"trade-{record['trade'].id}",
                "label": record["opened_at_local"].strftime("%d %b")
                if record["opened_at_local"]
                else "Unknown",
                "equity": round(running_equity, 2),
                "pnl": round(record["pnl"], 2),
                "symbol": record["symbol"],
            }
        )

    weekday_buckets = {
        name: {"label": name[:3], "count": 0, "wins": 0, "pnl": 0.0}
        for name in WEEKDAY_NAMES
    }
    pair_buckets = defaultdict(lambda: {"count": 0, "wins": 0, "pnl": 0.0, "gross_profit": 0.0, "gross_loss": 0.0})
    session_buckets = defaultdict(lambda: {"count": 0, "wins": 0, "pnl": 0.0})

    for record in closed_records:
        weekday_bucket = weekday_buckets.get(record["weekday"])
        if weekday_bucket is not None:
            weekday_bucket["count"] += 1
            weekday_bucket["pnl"] += record["pnl"]
            if record["pnl"] > 0:
                weekday_bucket["wins"] += 1

        pair_bucket = pair_buckets[record["symbol"] or "Unknown"]
        pair_bucket["count"] += 1
        pair_bucket["pnl"] += record["pnl"]
        if record["pnl"] > 0:
            pair_bucket["wins"] += 1
            pair_bucket["gross_profit"] += record["pnl"]
        elif record["pnl"] < 0:
            pair_bucket["gross_loss"] += record["pnl"]

        session_bucket = session_buckets[record["session"]]
        session_bucket["count"] += 1
        session_bucket["pnl"] += record["pnl"]
        if record["pnl"] > 0:
            session_bucket["wins"] += 1

    weekday_stats = []
    for name in WEEKDAY_NAMES:
        bucket = weekday_buckets[name]
        weekday_stats.append(
            {
                "name": name,
                "label": bucket["label"],
                "count": bucket["count"],
                "win_rate": (bucket["wins"] / bucket["count"] * 100.0) if bucket["count"] else None,
                "avg_pnl": (bucket["pnl"] / bucket["count"]) if bucket["count"] else None,
                "net_pnl": bucket["pnl"],
            }
        )

    pair_stats = []
    for symbol, bucket in pair_buckets.items():
        count = bucket["count"]
        pair_stats.append(
            {
                "symbol": symbol,
                "count": count,
                "win_rate": (bucket["wins"] / count * 100.0) if count else None,
                "expectancy": (bucket["pnl"] / count) if count else None,
                "net_pnl": bucket["pnl"],
                "profit_factor": (
                    bucket["gross_profit"] / abs(bucket["gross_loss"])
                    if bucket["gross_loss"] < 0
                    else None
                ),
            }
        )
    pair_stats.sort(key=lambda item: (-item["count"], -item["net_pnl"], item["symbol"]))

    session_stats = []
    for session_name, bucket in session_buckets.items():
        count = bucket["count"]
        session_stats.append(
            {
                "name": session_name,
                "count": count,
                "win_rate": (bucket["wins"] / count * 100.0) if count else None,
                "expectancy": (bucket["pnl"] / count) if count else None,
                "net_pnl": bucket["pnl"],
            }
        )
    session_stats.sort(key=lambda item: (-item["count"], item["name"]))

    best_trade = max(closed_records, key=lambda record: record["pnl"], default=None)
    worst_trade = min(closed_records, key=lambda record: record["pnl"], default=None)
    streaks = calculate_streaks(closed_records)

    return {
        "timezone_name": getattr(display_timezone, "key", "UTC"),
        "summary": {
            "total_trades": total_trades,
            "closed_trades": total_closed,
            "open_trades": open_trades,
            "wins": wins,
            "losses": losses,
            "breakeven": breakeven,
            "win_rate": win_rate,
            "net_pnl": net_pnl,
            "weekly_pnl": weekly_pnl,
            "monthly_pnl": monthly_pnl,
            "gross_profit": gross_profit,
            "gross_loss": gross_loss,
            "expectancy": expectancy,
            "avg_win": avg_win,
            "avg_loss_abs": avg_loss_abs,
            "payoff_ratio": payoff_ratio,
            "risk_reward_ratio": risk_reward_ratio,
            "profit_factor": profit_factor,
            "average_duration_minutes": average_duration_minutes,
            "average_duration_label": format_duration_minutes(average_duration_minutes),
            "duration_coverage": duration_coverage,
            "trading_days": len(trading_days),
            "max_drawdown": abs(max_drawdown),
            "best_trade": best_trade,
            "worst_trade": worst_trade,
        },
        "streaks": streaks,
        "equity_curve": equity_curve,
        "daily_equity_curve": list(daily_equity.values()),
        "weekday_stats": weekday_stats,
        "pair_stats": pair_stats,
        "session_stats": session_stats,
        "closed_records": closed_records,
        "week_label": week_start_local.strftime("%d %b %Y"),
    }
